from celery import shared_task
from django.utils import timezone
from io import BytesIO
import pandas as pd

from openpyxl.styles import Font

from .models import (
    SqlQuery,
    Report,
    ReportExecutionLog,
    ReportQueryParameter,
    ReportEmail,
)
from .utils import execute_sql_on_remote, send_report_email
from .utils import humanize_email_error


# =====================================================
# 🔹 UTILITAIRE : CALCUL DES TOTAUX (SAFE)
# =====================================================
def compute_totals(df, columns):
    """
    Calcule les totaux des colonnes numériques demandées
    sans jamais casser l'exécution.
    """
    totals = {}

    for col in columns:
        if col in df.columns:
            try:
                totals[col] = pd.to_numeric(
                    df[col],
                    errors="coerce"
                ).fillna(0).sum()
            except Exception:
                totals[col] = 0

    return totals


# =====================================================
# 🔹 Exécution d'une requête SQL simple
# =====================================================
@shared_task(
    bind=True,
    autoretry_for=(Exception,),
    retry_kwargs={
        "countdown": 1800,  # 30 minutes
        "max_retries": 3,
    },
)
def execute_sql_query_task(self, query_id):
    query = SqlQuery.objects.get(id=query_id)

    try:
        df = execute_sql_on_remote(
            query.database,
            query.sql_text
        )

        if df.empty:
            raise ValueError("La requête SQL n'a retourné aucune donnée.")

        # ==========================
        # TOTAUX (REQUÊTE SEULE)
        # ==========================
        enable_totals = getattr(query, "enable_totals", False)
        total_columns = getattr(query, "total_columns", [])

        if enable_totals and total_columns:
            totals = compute_totals(df, total_columns)

            if totals:
                total_row = {col: "" for col in df.columns}
                label_col = df.columns[0]
                total_row[label_col] = getattr(query, "total_label", "TOTAL")

                for col, value in totals.items():
                    total_row[col] = value

                df = pd.concat(
                    [df, pd.DataFrame([total_row])],
                    ignore_index=True
                )

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Résultats")

            # Style TOTAL
            if enable_totals and not df.empty:
                ws = writer.book["Résultats"]
                last_row = ws.max_row
                for cell in ws[last_row]:
                    cell.font = Font(bold=True)

        output.seek(0)

        raise ValueError(
            "Aucun destinataire TO défini pour l’exécution de requête seule."
        )

    except Exception as e:
        ReportExecutionLog.objects.create(
            report=None,
            query=query,
            status="error",
            message=str(e),
        )
        raise


# =====================================================
# 🔹 Exécution d’un RAPPORT (UN EXCEL / PLUSIEURS FEUILLES)
# =====================================================
@shared_task(
    bind=True,
    autoretry_for=(Exception,),
    retry_kwargs={
        "countdown": 1800,  # 30 minutes
        "max_retries": 3,
    },
)
def execute_report_task(self, report_id):

    report = Report.objects.get(id=report_id)

    ReportExecutionLog.objects.create(
        report=report,
        query=None,
        status="success",
        message="Démarrage de l’exécution du rapport",
    )

    if not report.queries.exists():
        ReportExecutionLog.objects.create(
            report=report,
            query=None,
            status="error",
            message="Aucune requête associée au rapport",
        )
        return

    has_error = False
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        for index, query in enumerate(report.queries.all(), start=1):
            try:
                params = {
                    p.name: p.value
                    for p in ReportQueryParameter.objects.filter(
                        report=report,
                        query=query
                    )
                }

                df = execute_sql_on_remote(
                    query.database,
                    query.sql_text,
                    params
                )

                if df.empty:
                    df = pd.DataFrame({
                        "INFO": [f"Aucune donnée retournée pour la requête : {query.name}"]
                    })

                # ==========================
                # TOTAUX (PAR REQUÊTE)
                # ==========================
                enable_totals = getattr(query, "enable_totals", False)
                total_columns = getattr(query, "total_columns", [])

                if enable_totals and total_columns:
                    totals = compute_totals(df, total_columns)

                    if totals:
                        total_row = {col: "" for col in df.columns}
                        label_col = df.columns[0]
                        total_row[label_col] = getattr(query, "total_label", "TOTAL")

                        for col, value in totals.items():
                            total_row[col] = value

                        df = pd.concat(
                            [df, pd.DataFrame([total_row])],
                            ignore_index=True
                        )

                sheet_name = f"{index}_{query.name}"[:31]

                df.to_excel(
                    writer,
                    index=False,
                    sheet_name=sheet_name
                )

                # Style TOTAL
                if enable_totals and not df.empty:
                    ws = writer.book[sheet_name]
                    last_row = ws.max_row
                    for cell in ws[last_row]:
                        cell.font = Font(bold=True)

                ReportExecutionLog.objects.create(
                    report=report,
                    query=query,
                    status="success",
                    message=f"Requête exécutée avec succès ({sheet_name})",
                )

            except Exception as e:
                has_error = True

                ReportExecutionLog.objects.create(
                    report=report,
                    query=query,
                    status="error",
                    message=str(e),
                )

                error_df = pd.DataFrame({"ERREUR": [str(e)]})
                error_df.to_excel(
                    writer,
                    sheet_name=f"ERR_{index}"[:31],
                    index=False
                )

    output.seek(0)

    attachments = [{
        "filename": f"{report.code}_{report.name}.xlsx",
        "content": output.getvalue(),
        "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }]

    # =================================================
    # 📧 EMAIL
    # =================================================
    to_emails = list(
        ReportEmail.objects.filter(
            report=report,
            email_type="to"
        ).values_list("email", flat=True)
    )

    cc_emails = list(
        ReportEmail.objects.filter(
            report=report,
            email_type="cc"
        ).values_list("email", flat=True)
    )

    if not to_emails:
        has_error = True
        ReportExecutionLog.objects.create(
            report=report,
            query=None,
            status="error",
            message="Aucun destinataire TO défini pour le rapport",
        )
    else:
        try:
            send_report_email(
                subject=report.subject or f"Rapport : {report.name}",
                body=report.message or "Veuillez trouver le rapport en pièce jointe.",
                to_emails=to_emails,
                cc_emails=cc_emails,
                attachments=attachments,
            )
            ReportExecutionLog.objects.create(
                report=report,
                query=None,
                status="success",
                message="Rapport envoyé par email avec succès",
            )

        except Exception as e:
            has_error = True
            ReportExecutionLog.objects.create(
                report=report,
                query=None,
                status="error",
                message=humanize_email_error(e),
            )
            raise

    report.last_executed_at = timezone.now()
    report.save(update_fields=["last_executed_at"])

    return (
        f"Rapport '{report.name}' exécuté avec erreurs"
        if has_error
        else f"Rapport '{report.name}' exécuté avec succès"
    )
